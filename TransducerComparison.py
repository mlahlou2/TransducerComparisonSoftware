import time
#import ticexec as tc1
#import msvcrt
import decimal
import datetime
import openpyxl as pyxl

import tkinter as tk
from tkinter import *
from tkinter import ttk
import tkinter.messagebox
import tkinter.filedialog as filedialog
import tkinter.simpledialog as sdialog
import tkinter.messagebox as msgbox
import pandas as pd
import scipy.stats as stats

import redpitaya_scpi as rp
from matplotlib.backends.backend_pdf import PdfPages
import matplotlib.pyplot as plt
import hyd_calibration
import calibrationsheet
import hydrophone as hyd
import numpy as np

import os.path
import datetime
import fpdf

from hyd_calibration import hyd_calibration_multiple_freq

# Writing to an excel
# sheet using Python
import xlwt
from xlwt import Workbook


AMPLIFIER = 'A150_'
LARGE_FONT= ("Verdana", 12)
NORM_FONT = ("Helvetica", 10)
SMALL_FONT = ("Helvetica", 8)


def popupmsg(msg):
    popup = tk.Tk()
    popup.wm_title("!")
    label = ttk.Label(popup, text=msg, font=NORM_FONT)
    label.pack(side="top", fill="x", pady=10)
    B1 = ttk.Button(popup, text="Okay", command=popup.destroy)
    B1.pack()
    # popup.mainloop()

class VoltSweepGUI:
    """ Create a GUI for voltage sweep """
    def __init__(self,master):
        self.createWidgets()

    def createWidgets(self):
        self.frame = Frame()
        self.frame.pack()

        self.b1 = Button(self.frame, text="Compare A Transducer", command=self.comparetransducer).grid(row=1,
                                                                                        column=1, padx=20, columnspan=4)
        self.b2 = Button(self.frame, text="Compare A Group of Transducers", command=self.comparealltransducers).grid(row=3,
                                                                                        column=1, padx=20, columnspan=2)
        self.integer_entry_state = IntVar()
        integer_entry_text = 'Transducer to compare'
        self.integer_entrybox = Entry(self.frame, textvariable=self.integer_entry_state).grid(row=3,
                                                                                        column=3, padx=20, columnspan=1)
        self.integer_Label = Label(self.frame, text=integer_entry_text).grid(row=3, column=4, padx=20, columnspan=1)
        self.b3 = Button(self.frame, text="Compare A Group of Transducers Beamprofiles", command=self.comparealltransducers_beamprofile).grid(row=4,
                                                                                        column=1, padx=20, columnspan=2)
        self.b4 = Button(self.frame, text="Compare A Group of Transducers Info", command=self.comparealltransducers_All).grid(row=5,
                                                                                        column=1, padx=20, columnspan=2)

    def vnoamp(self, keypressures, vs):
        # calculating the voltage needed for the pnp values desired
        vnoamp = vs.voltage * 2e3
        vnoamp_round = np.round(vnoamp, decimals=3)
        vamp = vs.voltage * 2 * (10 ** (vs.amplify / 20))
        vamp_round = np.round(vamp, decimals=3)
        # input of hyd_calibration should be in MHz
        sensitivity = hyd_calibration.hyd_calibration(vs.cfreq)
        # calculating pnp
        pnp = -1e-6 * np.min(vs.hydoutput, axis=1) / sensitivity
        pnp_round = np.round(pnp, decimals=3)
        # calculating MI based off of the volt sweep data
        mi = pnp / np.sqrt(vs.cfreq)
        for i in range(len(mi)):
            mi[i] = np.format_float_scientific(mi[i], precision=3)
        table_data2 = [vnoamp_round, vamp_round, pnp_round, mi]
        table_data2 = np.transpose(table_data2)

        cal325LA = msgbox.askquestion('Calibrated with 325LA?', 'Calibrated with 325LA?') == 'yes'
        if cal325LA:
            voltscale = 1.77
        else:
            voltscale = 1.0

        # interpolated data for key pressures
        couplerbool = msgbox.askquestion('Calibrated with couplers?', 'Calibrated with couplers?') == 'yes'
        kp_noamp = np.interp(keypressures, pnp, vnoamp / voltscale)
        # kp_amp = np.interp(keypressures, pnp, vamp/voltscale)
        couplvalue = -0.5375  # in db for pair of directional couplers from minicircuits
        if couplerbool:
            kp_noamp_coupl = kp_noamp
            kp_noamp = kp_noamp * (10 ** (couplvalue / 20))  # account for attenuation
        else:
            kp_noamp_coupl = kp_noamp / (10 ** (couplvalue / 20))  # account for attenuation

        return np.round(kp_noamp_coupl, decimals=2), table_data2[-1][1:-1]

    def loadfiles(self, freq_data, beam_data, vs, fs, bp, vs_file, fs_file, bp_file):
        if vs_file == '':
            vs_file = filedialog.askopenfilename(title='Select Volt Sweep File')
            vs.load(vs_file)
        if freq_data == True:
            if fs_file == '':
                fs_file = filedialog.askopenfilename(title='Select Frequency Sweep File')
            if fs_file == '':
                freq_data = False
            else:
                fs.load(fs_file)
        if beam_data == True:
            if bp_file == '':
                bp_file = filedialog.askopenfilename(title='Select Beam Profile File')
            bp.load(bp_file)

        if vs.amplify == 0:
            vs.amplify = float(sdialog.askstring('Amplification', 'Amplification is set to 0 dB. Please enter amplification in dB'))
        return vs, fs, bp, vs_file, fs_file, bp_file

    def loadbpfiles(self, freq_data, beam_data, vs, fs, bp, vs_file, fs_file, bp_file):
        if beam_data == True:
            if bp_file == '':
                bp_file = filedialog.askopenfilename(title='Select Beam Profile File')
            bp.load(bp_file)
        return bp, bp_file

    def plot_transducer_scatter(self, workbook_ID, transducer_ID_sheet_name, vs_file_current, fs_file_current, vs_file_prev, fs_file_prev):
        ##############################################################################
        #
        # An example of creating a chart with Pandas and XlsxWriter.
        #

        # Create a Pandas dataframe from the data.
        index_vs = ['Current VS','Previous VS']
        df = pd.DataFrame([vs_file_current, vs_file_prev], index=['Current VS', 'Previous VS'])
        df = df.reindex(columns=sorted(df.columns))
        print(df)

        # Create a Pandas Excel writer using XlsxWriter as the engine.
        excel_file = workbook_ID
        sheet_name = transducer_ID_sheet_name

        writer = pd.ExcelWriter(excel_file, engine='xlsxwriter')
        df.to_excel(writer, sheet_name=sheet_name)

        # Access the XlsxWriter workbook and worksheet objects from the dataframe.
        workbook = writer.book
        worksheet = writer.sheets[sheet_name]

        # # Create a chart object.
        # vs_chart = workbook.add_chart({'type': 'scatter',
        #                             'subtype': 'smooth'})
        #
        # # Configure the series of the chart from the dataframe data.
        # max_row = len(df)
        # for i in range(len()):
        #     col = i + 1
        #     chart.add_series({
        #         'name': ['Sheet1', 0, col],
        #         'categories': ['Sheet1', 1, 0, max_row, 0],
        #         'values': ['Sheet1', 1, col, max_row, col],
        #         'marker': {'type': 'circle', 'size': 7},
        #     })
        #
        # # Configure the chart axes.
        # chart.set_x_axis({'name': 'Index'})
        # chart.set_y_axis({'name': 'Data Value',
        #                   'major_gridlines': {'visible': False}})
        #
        # # Insert the chart into the worksheet.
        # worksheet.insert_chart('K2', chart)
        #
        # # Close the Pandas Excel writer and output the Excel file.
        # writer.save()

    def get_time_of_measurements(self, vs_file):
        # getting the date the volt sweep file was created
        created = os.path.getctime(vs_file)
        date = datetime.datetime.fromtimestamp(created)
        return date.date()

    def closout_pdf(self,pdf, pdfname, plt):
        pdf.output(pdfname)
        pdf.close()
        plt.close('all')

    def comparetransducer(self, vs_current_file='', fs_current_file='',bp_current_file='',  vs_previous_file='', fs_previous_file='',bp_previous_file='', transducername= '',freq_data= True, beam_data=True):
        '''Generates a pdf report from voltsweep, freqsweep and beam profile data
            INPUTS:
                vs_file: voltsweep data file path, string (if no path is provided user will be prompted to select file from directory)
                fs_file: freqsweep data file path, string (if no path is provided user will be prompted to select file from directory)
                bp_file: beamprofile data file path  string (if no path is provided user will be prompted to select file from directory)
                pdfname: path of pdf to be saved, string (if no path is provided user will be prompted to select save location and name)
                freq_data: boolean, if True, the user has freq sweep data to be included in the report, if False no freq sweep
                           data will be included in the report
                beam_data: boolean, if True, the user has beam profile data to be included in the report, if False no beam profile
                           data will be included in the report
            OUTPUT:
                pdf document saved as directed by user
                '''


        # creating objects
        vs_current = hyd.VoltSweep()
        fs_current = hyd.FreqSweep()
        bp_current = hyd.BeamProfile()
        vs_previous = hyd.VoltSweep()
        fs_previous = hyd.FreqSweep()
        bp_previous = hyd.BeamProfile()

        # Loading Current Transducer Calibration Data
        [vs_current, fs_current, bp_current, vs_current_file, fs_current_file, bp_current_file] = self.loadfiles(freq_data, beam_data, vs_current, fs_current, bp_current, vs_current_file, fs_current_file,
                       bp_current_file)

        # Loading Previous Transducer Calibration Data
        [vs_previous, fs_previous, bp_previous, vs_previous_file, fs_previous_file, bp_previous_file] = self.loadfiles(freq_data, beam_data, vs_previous, fs_previous, bp_previous, vs_previous_file, fs_previous_file,
                       bp_previous_file)

        ### Voltage Table Creation ###
        # collects the transducers number from the current voltsweep data
        if transducername == '':
            transducername = vs_current.txdr

        current_date = self.get_time_of_measurements(vs_current_file)
        previous_date = self.get_time_of_measurements(vs_previous_file)

        ### Voltageoutputs and PNP for the largest voltage input
        keypressures = np.array([0.3, 0.5, 1.5])  # in MPa
        [current_voltageoutput, current_max_pnp] = self.vnoamp(keypressures, vs_current)
        [previous_voltageoutput, previous_max_pnp] = self.vnoamp(keypressures, vs_previous)

        ### PNP table
        collabels = ['Transducer', 'Vamp (Vpp)', 'Cur PNP(MPa)', 'Prev PNP(MPa)', "% Change"]
        percentagePNPchange = (current_max_pnp[1]-previous_max_pnp[1])/previous_max_pnp[1]
        percentagePNPstring = str(np.round((100.0 * percentagePNPchange), decimals=2)) + '%'
        table_PNP = [transducername, current_max_pnp[0], current_max_pnp[1], previous_max_pnp[1], percentagePNPstring]
        table_PNP = np.transpose(table_PNP)
        print(table_PNP)
        # ###Voltage Outputs
        # percentagevoltageoutputchange = (current_voltageoutput-previous_voltageoutput)/previous_voltageoutput
        # percentagevoltageoutputstring = []
        # for i, j in zip(range(len(percentagevoltageoutputchange)), percentagevoltageoutputchange):
        #     percentagevoltageoutputstring.append(str(np.round((100.0 * j), decimals=2)) + '%')
        # table_data_kp = [keypressures, np.round(keypressures / np.sqrt(vs_current.cfreq), decimals=3), previous_voltageoutput, current_voltageoutput, percentagevoltageoutputstring]
        # table_data_kp = np.transpose(table_data_kp)
        # collabels_kp = ['PNP (MPa)', 'MI', 'Prev Vnoamp(Vpp)', 'Cur Vnoamp(Vpp)', '% Change']
        # print(table_data_kp)

        pdf = fpdf.FPDF()
        pdf.add_page()
        pdf.set_font('Arial','',12)
        if (vs_current.txdr == ''):
            vs_current.txdr = sdialog.askstring('Transducer Name', 'Please enter transducer name')
        pdf.cell(0, 6, 'Transducer: ' + vs_current.txdr, ln=1)
        pdf.cell(0, 6, 'Current Measurement made: ' + str(current_date), ln=1)
        pdf.cell(0, 6, 'Previous Measurement made: ' + str(previous_date), ln=1)
        # calibrationsheet.BasicFPDFTable(pdf, table_data_kp, collabels_kp)
        calibrationsheet.BasicFPDFTable(pdf, table_PNP, collabels)

        pdfname = 'C:/Users/SoundPipe/Desktop/' + str(current_date) + '_Transducer_' + transducername + '.pdf'
        self.closout_pdf(pdf, pdfname, plt)



    #     try:
    #         beamwidth = bp.beamwidth()
    #     except:
    #         beamwidth = 0
    #     # creating table of settings/info about the transducer, add rows and columns to include more information.
    #     table_data = [
    #         ["Txdr #", vs.txdr],
    #         ["Matching Network:", vs.matchingnetwork],
    #         ["Amplifier", vs.amplifier],
    #         ["Frequency", vs.cfreq.item()],
    #         ["-6 dB Beamwidth", beamwidth],
    #         ["Date", date2]
    #
    #     ]
    #
    #     table = ax.table(cellText=table_data, loc='center')
    #     table.set_fontsize(14)
    #     table.scale(1, 4)
    #     ax.axis('off')
    #     # saves the figure to one page of the pdf document
    #     # pp.savefig()
    #
    #     # adding subplots for volt sweep, freq sweep, beam profile and the PNP/MI table
    #     fig = plt.figure()
    #
    #     # these functions are necessary to make the secondary MI axis for the voltsweep plot
    #     def pnp2mi(pnp):
    #         return pnp / np.sqrt(vs.cfreq)
    #
    #     def mi2pnp(pnp):
    #         return pnp * np.sqrt(vs.cfreq)
    #
    #     # plotting volt sweep data
    #     [vs_figure1, vs_figpath] = vs.plot(displayplt=False, saveplt=True,
    #                                        savepath=os.path.splitext(vs_file)[0] + '.png')
    #     # axa = fig.add_subplot(2, 2, 1)
    #     # sensitivity = hyd_calibration.hyd_calibration(vs.cfreq)
    #     # pnp = -1e-6 * np.min(vs.hydoutput, axis=1) / sensitivity
    #     x = vs.voltage * np.power(10.0, (vs.amplify / 20.0))
    #     # figure3 = axa.plot(x, pnp, 'x')
    #     # axa.set_xlabel('Input Voltage (v)')
    #     # axa.set_ylabel('Peak Negative Pressure (MPa)')
    #     # axa.set_title('Voltage Sweep')
    #     # ax2 = axa.secondary_yaxis(location='right', functions=(pnp2mi, mi2pnp))
    #     # ax2.set_ylabel('MI')
    #     # ax2.ticklabel_format(axis="y", style="sci", scilimits=(0, 0))
    #     # plotting freq sweep data
    #     if freq_data == True:
    #         #    axb = fig.add_subplot(2, 2, 2)
    #         #    sensitivity = hyd_calibration.hyd_calibration_multiple_freq(fs.cfreq)
    #         #    pnp = -1e-6 * np.min(fs.hydoutput, axis=1) / sensitivity
    #         #    figure2 = axb.plot(fs.cfreq, pnp, 'x')
    #         #    axb.set_title('Frequency Sweep')
    #         #    plt.xlabel('Frequency (MHz)')
    #         #    plt.ylabel('Peak Negative Pressure (MPa)')
    #         #    axe = fig.add_subplot(2,2,4)
    #         #    mi_fs = pnp/np.sqrt(fs.cfreq)
    #         #    figure4 = axe.plot(fs.cfreq, mi_fs, 'x')
    #         #    axe.set_title('Frequency Sweep')
    #         #    plt.xlabel('Frequency (MHz)')
    #         #    plt.ylabel('MI')
    #         [fs_figure, fs_figpath] = fs.plot(displayplt=False, saveplt=True,
    #                                           savepath=os.path.splitext(fs_file)[0] + '.png')
    #     # plotting beam profile data
    #     if beam_data == True:
    #         [bp_figure, bp_figpath] = bp.plot(displayplt=False, saveplt=True,
    #                                           savepath=os.path.splitext(bp_file)[0] + '.png')
    #         # axc = fig.add_subplot(223, projection='polar')
    #         # axc.set_title('Beam Profile, PNP (MPa)', pad=15)
    #         # axc.plot(bp.angle * np.pi / 180.0, bp.hydoutput * 1e-6) #TODO  - correct to pressure
    #
    #     # expanding the subplot window so that the plots are more spread out and bigger on the pdf
    #     manager = plt.get_current_fig_manager()
    #     manager.resize(*manager.window.maxsize())
    #     plt.tight_layout()
    #     # plt.show()
    #     # pp.savefig(fig)
    #
    #     # TODO check this for accuracy
    #     # Creating information table and putting it on a separate pdf page
    #     fig2 = plt.figure()
    #     axd = fig2.add_subplot(1, 1, 1)
    #     vnoamp = vs.voltage * 2e3
    #     vnoamp_round = np.round(vnoamp, decimals=3)
    #     vamp = vs.voltage * 2 * (10 ** (vs.amplify / 20))
    #     vamp_round = np.round(vamp, decimals=3)
    #     # input of hyd_calibration should be in MHz
    #     sensitivity = hyd_calibration.hyd_calibration(vs.cfreq)
    #     # calculating pnp
    #     pnp = -1e-6 * np.min(vs.hydoutput, axis=1) / sensitivity
    #     pnp_round = np.round(pnp, decimals=3)
    #     # calculating MI based off of the volt sweep data
    #     mi = pnp / np.sqrt(vs.cfreq)
    #     for i in range(len(mi)):
    #         mi[i] = np.format_float_scientific(mi[i], precision=3)
    #     table_data2 = [vnoamp_round, vamp_round, pnp_round, mi]
    #     table_data2 = np.transpose(table_data2)
    #     collabels = ['Vnoamp (mVpp)', 'Vamp (Vpp)', 'PNP (MPa)', 'MI']
    #     # table = axd.table(cellText=table_data2, colLabels=collabels, loc='center', bbox=[0.05, 0.05, 0.8, 0.6])
    #     # table.set_fontsize(8)
    #     # table.scale(1, 4)
    #     # axd.axis('off')
    #     # plt.show()
    #     # pp.savefig()
    #     # pp.close() #call this after all of the figures are saved
    #
    #     cal325LA = msgbox.askquestion('Calibrated with 325LA?', 'Calibrated with 325LA?') == 'yes'
    #     if cal325LA:
    #         voltscale = 1.77
    #     else:
    #         voltscale = 1.0
    #
    #     # interpolated data for key pressures
    #     couplerbool = msgbox.askquestion('Calibrated with couplers?', 'Calibrated with couplers?') == 'yes'
    #     keypressures = np.array([0.3, 0.5, 1.5])  # in MPa
    #     kp_noamp = np.interp(keypressures, pnp, vnoamp / voltscale)
    #     # kp_amp = np.interp(keypressures, pnp, vamp/voltscale)
    #     couplvalue = -0.5375  # in db for pair of directional couplers from minicircuits
    #     if couplerbool:
    #         kp_noamp_coupl = kp_noamp
    #         kp_noamp = kp_noamp * (10 ** (couplvalue / 20))  # account for attenuation
    #     else:
    #         kp_noamp_coupl = kp_noamp / (10 ** (couplvalue / 20))  # account for attenuation
    #
    #     table_data_kp = [keypressures, np.round(keypressures / np.sqrt(vs.cfreq), decimals=3),
    #                      np.round(kp_noamp, decimals=1), np.round(kp_noamp_coupl, decimals=1)]
    #     table_data_kp = np.transpose(table_data_kp)
    #     # table_data_kp.append(['', '', '', ''])
    #     collabels_kp = ['PNP (MPa)', 'MI', 'Vnoamp w/o couplers (mVpp)', 'Vnoamp w/ couplers (Vpp)']
    #
    #     pdf = fpdf.FPDF()
    #     pdf.add_page()
    #     pdf.set_font('Arial', '', 12)
    #     if (vs.txdr == ''):
    #         vs.txdr = sdialog.askstring('Transducer Name', 'Please enter transducer name')
    #     pdf.cell(0, 6, 'Transducer: ' + vs.txdr, ln=1)
    #     pdf.cell(0, 6, 'Amplifier: ' + vs.amplifier, ln=1)
    #     pdf.cell(0, 6, 'Matching Network: ' + vs.matchingnetwork, ln=1)
    #     pdf.cell(0, 6, 'Frequency: ' + str(vs.cfreq.item()), ln=1)
    #     pdf.cell(0, 6, "-6 dB Beamwidth: " + str(beamwidth), ln=1)
    #     # pdf.cell(0, 10, 'Distance from Hydrophone: ' + str(vs.depth.item()), ln=1)
    #     # pdf.cell(0, 10, 'Collection Date: ' + vs.collectiondate , ln=1)
    #
    #     curline = pdf.get_y()
    #     pdf.image(vs_figpath, w=100)
    #     pdf.image(fs_figpath, w=100, y=curline, x=110)
    #     pdf.image(bp_figpath, w=100, x=55)
    #     BasicFPDFTable(pdf, table_data_kp, collabels_kp)
    #     BasicFPDFTable(pdf, table_data2, collabels)
    #
    #     pdf.output(pdfname)
    #     pdf.close()
    #     plt.close('all')
    #
    #     print("compare a transducer")
    #
    def comparealltransducers(self, vs_current_file='', fs_current_file='',bp_current_file='',  vs_previous_file='', fs_previous_file='',bp_previous_file='', transducername= '',freq_data= True, beam_data=True):

        '''Generates a pdf report from voltsweep, freqsweep and beam profile data
            INPUTS:
                vs_file: voltsweep data file path, string (if no path is provided user will be prompted to select file from directory)
                fs_file: freqsweep data file path, string (if no path is provided user will be prompted to select file from directory)
                bp_file: beamprofile data file path  string (if no path is provided user will be prompted to select file from directory)
                pdfname: path of pdf to be saved, string (if no path is provided user will be prompted to select save location and name)
                freq_data: boolean, if True, the user has freq sweep data to be included in the report, if False no freq sweep
                           data will be included in the report
                beam_data: boolean, if True, the user has beam profile data to be included in the report, if False no beam profile
                           data will be included in the report
            OUTPUT:
                excel document saved as directed by user
                '''

        pd.set_option("display.max_rows", None, "display.max_columns", None)
        if self.integer_entry_state.get() == 0:
            print('0 is not a valid entry')
            return

        # Spreadsheet_Name = ('C:/Users/SoundPipe/Desktop/MutliTransducerComparisonFor' +str(self.integer_entry_state.get()) + '.xlsx')
        Spreadsheet_Name = ('MutliTransducerComparisonFor' +str(self.integer_entry_state.get()) + '.xlsx')
        # Workbook is created
        excel_spreadsheet_writer = pd.ExcelWriter(Spreadsheet_Name, engine='xlsxwriter')

        collabels = ['Transducer', 'Current Measurement Date', 'Previous Measurement Date', 'Vamp (Vpp)', 'Cur PNP(MPa)', 'Prev PNP(MPa)', "% Change"]

        TransducerData = pd.DataFrame([], columns=collabels)

        for i in range(1, self.integer_entry_state.get()+1):

            # creating objects
            vs_current = hyd.VoltSweep()
            fs_current = hyd.FreqSweep()
            bp_current = hyd.BeamProfile()
            vs_previous = hyd.VoltSweep()
            fs_previous = hyd.FreqSweep()
            bp_previous = hyd.BeamProfile()
            vs_current_file = ''
            fs_current_file = ''
            bp_current_file = ''
            vs_previous_file = ''
            fs_previous_file = ''
            bp_previous_file = ''
            transducername = ''

            # Loading Current Transducer Calibration Data
            [vs_current, fs_current, bp_current, vs_current_file, fs_current_file, bp_current_file] = self.loadfiles(
                freq_data, beam_data, vs_current, fs_current, bp_current, vs_current_file, fs_current_file,
                bp_current_file)

            # Loading Previous Transducer Calibration Data
            [vs_previous, fs_previous, bp_previous, vs_previous_file, fs_previous_file, bp_previous_file] = self.loadfiles(
                freq_data, beam_data, vs_previous, fs_previous, bp_previous, vs_previous_file, fs_previous_file,
                bp_previous_file)

            ### Voltage Table Creation ###
            # collects the transducers number from the current voltsweep data
            # if transducername == '':
            transducername = vs_current.txdr

            current_date = self.get_time_of_measurements(vs_current_file)
            previous_date = self.get_time_of_measurements(vs_previous_file)
            print(previous_date)

            ### Voltageoutputs and PNP for the largest voltage input
            keypressures = np.array([0.3, 0.5, 1.5])  # in MPa
            [current_voltageoutput, current_max_pnp] = self.vnoamp(keypressures, vs_current)
            [previous_voltageoutput, previous_max_pnp] = self.vnoamp(keypressures, vs_previous)

            ### PNP table
            percentagePNPchange = (current_max_pnp[1] - previous_max_pnp[1]) / previous_max_pnp[1]
            percentagePNPstring = str(np.round((100.0 * percentagePNPchange), decimals=2)) + '%'
            table_PNP = [transducername, str(current_date), str(previous_date), current_max_pnp[0], current_max_pnp[1], previous_max_pnp[1], percentagePNPstring]

            numpy_Table_PNP = np.array(table_PNP)
            table_PNPTransducerData = pd.DataFrame(numpy_Table_PNP.reshape(1, -1), columns=list(TransducerData))
            TransducerData = TransducerData.append(table_PNPTransducerData)
            # self.plot_transducer_scatter(workbook_ID=)

            # ###Goes through and calculate the pnp pressure and puts it in a pdf table
            # percentagevoltageoutputchange = (current_voltageoutput-previous_voltageoutput)/previous_voltageoutput
            # percentagevoltageoutputstring = []
            # for i, j in zip(range(len(percentagevoltageoutputchange)), percentagevoltageoutputchange):
            #     percentagevoltageoutputstring.append(str(np.round((100.0 * j), decimals=2)) + '%')
            # table_data_kp = [keypressures, np.round(keypressures / np.sqrt(vs_current.cfreq), decimals=3), previous_voltageoutput, current_voltageoutput, percentagevoltageoutputstring]
            # table_data_kp = np.transpose(table_data_kp)
            # collabels_kp = ['PNP (MPa)', 'MI', 'Prev Vnoamp(Vpp)', 'Cur Vnoamp(Vpp)', '% Change']
            # calibrationsheet.BasicPandasDataFrameTable(TransducerData, table_PNP, collabels)

        # Saves the Transducers data to the excel spreadsheet
        print(TransducerData)

        TransducerData.to_excel(excel_spreadsheet_writer, sheet_name='All Transducer Data', index=False)
        excel_spreadsheet_writer.save()

        ## Saves the pdf created and closes it out
        # pdfname = 'C:/Users/SoundPipe/Desktop/MutliTransducerComparisonFor' + str(current_date) + '.pdf'
        # self.closout_pdf(pdf, pdfname, plt)

        print("compare all transducers")

    def comparealltransducers_beamprofile(self, vs_current_file='', fs_current_file='',bp_current_file='',  vs_previous_file='', fs_previous_file='',bp_previous_file='', transducername= '',freq_data= True, beam_data=True):

        '''Generates a pdf report from voltsweep, freqsweep and beam profile data
            INPUTS:
                vs_file: voltsweep data file path, string (if no path is provided user will be prompted to select file from directory)
                fs_file: freqsweep data file path, string (if no path is provided user will be prompted to select file from directory)
                bp_file: beamprofile data file path  string (if no path is provided user will be prompted to select file from directory)
                pdfname: path of pdf to be saved, string (if no path is provided user will be prompted to select save location and name)
                freq_data: boolean, if True, the user has freq sweep data to be included in the report, if False no freq sweep
                           data will be included in the report
                beam_data: boolean, if True, the user has beam profile data to be included in the report, if False no beam profile
                           data will be included in the report
            OUTPUT:
                excel document saved as directed by user
                '''

        pd.set_option("display.max_rows", None, "display.max_columns", None)
        if self.integer_entry_state.get() == 0:
            print('0 is not a valid entry')
            return

        # Spreadsheet_Name = ('C:/Users/SoundPipe/Desktop/MutliTransducerComparisonFor' +str(self.integer_entry_state.get()) + '.xlsx')
        Spreadsheet_Name = ('MutliTransducerComparisonFor' +str(self.integer_entry_state.get()) + '.xlsx')
        # Workbook is created
        excel_spreadsheet_writer = pd.ExcelWriter(Spreadsheet_Name, engine='xlsxwriter')

        collabels = ['Transducer', 'Current Measurement Date', 'COV', 'degrees of beamprofile > 75% max pressure', 'average for > 75%', 'degrees of beamprofile > 50% max pressure', 'average for > 50%', 'Total Power Normalized']
        # for i, j in zip(collabels, range(len(collabels))):
        #     sheet1.write(0, j, i)

        TransducerData = pd.DataFrame([], columns=collabels)

        for i in range(1, self.integer_entry_state.get()+1):

            # creating objects
            vs_current = hyd.VoltSweep()
            fs_current = hyd.FreqSweep()
            bp_current = hyd.BeamProfile()
            vs_current_file = ''
            fs_current_file = ''
            bp_current_file = ''

            # Loading Current Transducer Calibration Data
            [bp_current, bp_current_file] = self.loadbpfiles(
                freq_data, beam_data, vs_current, fs_current, bp_current, vs_current_file, fs_current_file,
                bp_current_file)

            ### Voltage Table Creation ###
            # collects the transducers number from the current voltsweep data
            # if transducername == '':
            transducername = bp_current.txdr

            current_date = self.get_time_of_measurements(bp_current_file)



            bp_current_PNP = []
            for i in bp_current.hydoutput:
                bp_current_PNP.append(-min(i))

            COV = stats.variation(bp_current_PNP)
            TotalPower = sum(bp_current_PNP)
            bp_current_normalized = bp_current_PNP/max(bp_current_PNP)
            TotalPowerNormalized = sum(bp_current_normalized)
            average_output_nornalized = np.mean(bp_current_normalized)
            print(average_output_nornalized)
            high_pressure_75 = []
            high_pressure_50 = []
            low_pressure = []
            for i in bp_current_normalized:
                if i >= 0.75:
                    high_pressure_75.append(i)
                    high_pressure_50.append(i)
                elif i >= 0.5:
                    high_pressure_50.append(i)
                else:
                    low_pressure.append(i)

            degrees_above_threequarters = len(high_pressure_75)
            degrees_above_half = len(high_pressure_50)
            average_output_above_threequarters_normalized = np.mean(high_pressure_75)
            average_output_above_half_normalized = np.mean(high_pressure_50)
            table_PNP = [transducername, str(current_date), COV, degrees_above_threequarters, average_output_above_threequarters_normalized, degrees_above_half, average_output_above_half_normalized, TotalPowerNormalized]

            numpy_Table_PNP = np.array(table_PNP)
            table_PNPTransducerData = pd.DataFrame(numpy_Table_PNP.reshape(1, -1), columns=list(TransducerData))
            TransducerData = TransducerData.append(table_PNPTransducerData)

        # Saves the Transducers data to the excel spreadsheet
        print(TransducerData)

        TransducerData.to_excel(excel_spreadsheet_writer, sheet_name='All Transducer Data', index=False)
        excel_spreadsheet_writer.save()

        ## Saves the pdf created and closes it out
        # pdfname = 'C:/Users/SoundPipe/Desktop/MutliTransducerComparisonFor' + str(current_date) + '.pdf'
        # self.closout_pdf(pdf, pdfname, plt)

        print("compare all transducers")

    def comparealltransducers_All(self, vs_current_file='', fs_current_file='', bp_current_file='',
                                          vs_previous_file='', fs_previous_file='', bp_previous_file='',
                                          transducername='', freq_data=True, beam_data=True):

        '''Generates a pdf report from voltsweep, freqsweep and beam profile data
            INPUTS:
                vs_file: voltsweep data file path, string (if no path is provided user will be prompted to select file from directory)
                fs_file: freqsweep data file path, string (if no path is provided user will be prompted to select file from directory)
                bp_file: beamprofile data file path  string (if no path is provided user will be prompted to select file from directory)
                pdfname: path of pdf to be saved, string (if no path is provided user will be prompted to select save location and name)
                freq_data: boolean, if True, the user has freq sweep data to be included in the report, if False no freq sweep
                           data will be included in the report
                beam_data: boolean, if True, the user has beam profile data to be included in the report, if False no beam profile
                           data will be included in the report
            OUTPUT:
                excel document saved as directed by user
                '''

        pd.set_option("display.max_rows", None, "display.max_columns", None)
        if self.integer_entry_state.get() == 0:
            print('0 is not a valid entry')
            return

        # Spreadsheet_Name = ('C:/Users/SoundPipe/Desktop/MutliTransducerComparisonFor' +str(self.integer_entry_state.get()) + '.xlsx')
        Spreadsheet_Name = ('MutliTransducerComparisonFor' + str(self.integer_entry_state.get()) + '.xlsx')
        # Workbook is created
        excel_spreadsheet_writer = pd.ExcelWriter(Spreadsheet_Name, engine='xlsxwriter')

        collabels = ['Transducer', 'Current Measurement Date', 'Vamp (Vpp)', 'Cur PNP(MPa)', 'Peak Frequency', 'COV', 'degrees of beamprofile > 75% max pressure',
                     'average for > 75%', 'degrees of beamprofile > 50% max pressure', 'average for > 50%', 'Total Power of Transducer']

        TransducerData = pd.DataFrame([], columns=collabels)

        for i in range(1, self.integer_entry_state.get() + 1):

            # creating objects
            vs_current = hyd.VoltSweep()
            fs_current = hyd.FreqSweep()
            bp_current = hyd.BeamProfile()
            vs_current_file = ''
            fs_current_file = ''
            bp_current_file = ''

            # Loading Current Transducer Calibration Data
            [vs_current, fs_current, bp_current, vs_current_file, fs_current_file, bp_current_file] = self.loadfiles(
                freq_data, beam_data, vs_current, fs_current, bp_current, vs_current_file, fs_current_file,
                bp_current_file)

            ### Voltage Table Creation ###
            # collects the transducers number from the current voltsweep data
            # if transducername == '':
            transducername = bp_current.txdr
            current_date = self.get_time_of_measurements(bp_current_file)


        # goes through the beamprofile and creates a rating using cov and points 50% or great to the max value
            bp_current_PNP = []
            for i in bp_current.hydoutput:
                bp_current_PNP.append(-min(i))

            COV = stats.variation(bp_current_PNP)
            Totalpower = sum(bp_current_PNP)
            bp_current_normalized = bp_current_PNP / max(bp_current_PNP)
            TotalPowerNormalized = sum(bp_current_normalized)
            average_output_normalized = np.mean(bp_current_normalized)
            # print(average_output_normalized)
            high_pressure_75 = []
            high_pressure_50 = []
            low_pressure = []
            for i in bp_current_normalized:
                if i >= 0.75:
                    high_pressure_75.append(i)
                    high_pressure_50.append(i)
                elif i >= 0.5:
                    high_pressure_50.append(i)
                else:
                    low_pressure.append(i)

            degrees_above_threequarters = len(high_pressure_75)
            degrees_above_half = len(high_pressure_50)
            average_output_above_threequarters_normalized = np.mean(high_pressure_75)
            average_output_above_half_normalized = np.mean(high_pressure_50)

        # Goes thorugh and finds the voltage values to reach the current pnp settings for the transducer

            ### Voltageoutputs and PNP for the largest voltage input
            keypressures = np.array([0.3, 1.5])  # in MPa
            [current_voltageoutput, current_max_pnp] = self.vnoamp(keypressures, vs_current)

        # Goes through and get the peak center frequency
            sensitivity = hyd_calibration_multiple_freq(fs_current.cfreq)
            peakfreq =-1e-6 * np.min(fs_current.hydoutput,axis=1)/sensitivity
            indexmaxpeakfreq = np.argmax(peakfreq)
            maxpeakfreq = fs_current.cfreq[indexmaxpeakfreq]

        # Puts all the collect data in a table to be put in an excel spreadsheet
            table_PNP = [transducername, str(current_date), current_max_pnp[0], current_max_pnp[1], maxpeakfreq, COV, degrees_above_threequarters,
                         average_output_above_threequarters_normalized, degrees_above_half,
                         average_output_above_half_normalized, Totalpower]

            numpy_Table_PNP = np.array(table_PNP)
            table_PNPTransducerData = pd.DataFrame(numpy_Table_PNP.reshape(1, -1), columns=list(TransducerData))
            TransducerData = TransducerData.append(table_PNPTransducerData)

        # Saves the Transducers data to the excel spreadsheet
        # print(TransducerData)

        TransducerData.to_excel(excel_spreadsheet_writer, sheet_name='All Transducer Data', index=False)
        excel_spreadsheet_writer.save()

        ## Saves the pdf created and closes it out
        # pdfname = 'C:/Users/SoundPipe/Desktop/MutliTransducerComparisonFor' + str(current_date) + '.pdf'
        # self.closout_pdf(pdf, pdfname, plt)

        print("compare all transducers")
